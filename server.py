"""Веб-сервер интерактивного парсера инфраструктуры (FastAPI).

Запуск:
    .venv\\Scripts\\python.exe server.py
    # затем открыть http://127.0.0.1:8000

Фронтенд (static/index.html) обращается к /api/* — backend дёргает уже
готовую логику парсера (categories / Overpass / analytics / storage).
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

from io import BytesIO

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# UTF-8 в консоли Windows для логов
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8")
        except (ValueError, OSError):
            pass

from osm_parser import analytics, categories
from osm_parser.cities import city_list
from osm_parser.geo import polygon_area_km2
from osm_parser.geocode import GeocodeError, geocode, geocode_area
from osm_parser.sources import OverpassSource
from osm_parser.sources.base import PRESET_RADII_KM, Circle
from osm_parser.storage import Storage
from osm_parser.webmap import GROUP_COLORS

BASE = Path(__file__).parent
# Путь к БД можно переопределить (напр. на постоянный диск хостинга) через DB_PATH.
DB_PATH = os.environ.get("DB_PATH", str(BASE / "osm_poi.db"))
STATIC = BASE / "static"

app = FastAPI(title="pars-openstreetmap")
source = OverpassSource()


class SearchRequest(BaseModel):
    lat: float | None = None
    lon: float | None = None
    place: str | None = None
    radius_km: float = 2.0
    categories: list[str] | None = None
    name: str | None = None


class CitySearchRequest(BaseModel):
    city: str
    categories: list[str] | None = None
    name: str | None = None


class ShapeSearchRequest(BaseModel):
    geometry: dict          # GeoJSON Polygon (прямоугольник или произвольная фигура)
    categories: list[str] | None = None
    name: str | None = None


class ExportRequest(BaseModel):
    features: list[dict]
    titles: dict | None = None


class RenameRequest(BaseModel):
    name: str


def _summary_dict(summary) -> dict:
    return {
        "total": summary.total,
        "area_km2": summary.area_km2,
        "by_group": summary.by_group,
        "by_category": summary.by_category,
        "density": summary.density_per_km2,
    }


@app.get("/api/categories")
def api_categories() -> dict:
    return {
        "groups": categories.all_groups(),
        "categories": [
            {"key": c.key, "group": c.group, "title": c.title}
            for c in categories.all_categories()
        ],
        "presets": list(PRESET_RADII_KM),
        "groupColors": GROUP_COLORS,
    }


@app.get("/api/cities")
def api_cities() -> dict:
    return {"cities": city_list()}


@app.get("/api/geocode")
def api_geocode(q: str) -> dict:
    try:
        place = geocode(q)
    except GeocodeError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return {"lat": place.lat, "lon": place.lon, "display_name": place.display_name}


@app.post("/api/search")
def api_search(req: SearchRequest) -> dict:
    try:
        cats = categories.resolve(req.categories)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    # Определяем центр: координаты или геокодинг названия.
    display = None
    if req.lat is not None and req.lon is not None:
        lat, lon = req.lat, req.lon
    elif req.place:
        try:
            place = geocode(req.place)
        except GeocodeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        lat, lon, display = place.lat, place.lon, place.display_name
    else:
        raise HTTPException(status_code=400, detail="Нужны координаты или название места")

    if req.radius_km <= 0:
        raise HTTPException(status_code=400, detail="Радиус должен быть положительным")

    circle = Circle(lat, lon, req.radius_km * 1000.0)
    try:
        circle.validate()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    try:
        pois = list(source.fetch(cats, circle))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Ошибка источника данных: {exc}")

    summary = analytics.summarize(pois, circle.area_km2())
    rings = analytics.ring_analysis(pois, circle, rings=5)

    with Storage(DB_PATH) as store:
        store.save(pois)
        sid = store.record_search(
            place=req.place,
            center_lat=lat,
            center_lon=lon,
            radius_m=circle.radius_m,
            categories=[c.key for c in cats],
            result_count=len(pois),
            name=req.name,
        )

    return {
        "search_id": sid,
        "mode": "radius",
        "center": [lat, lon],
        "display_name": display,
        "radius_m": circle.radius_m,
        "count": len(pois),
        "features": [p.as_geojson_feature() for p in pois],
        "titles": {c.key: c.title for c in categories.all_categories()},
        "summary": _summary_dict(summary),
        "rings": [
            {"inner_km": r.inner_km, "outer_km": r.outer_km,
             "count": r.count, "density": r.density}
            for r in rings.rings
        ],
    }


@app.post("/api/search_area")
def api_search_area(req: CitySearchRequest) -> dict:
    try:
        cats = categories.resolve(req.categories)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    try:
        area = geocode_area(req.city)
    except GeocodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if area.geojson is None or area.osm_type not in ("relation", "way"):
        raise HTTPException(
            status_code=400,
            detail=f"У «{req.city}» нет административной границы (найдена точка, "
                   f"а не город/район). Уточните название.",
        )

    try:
        pois = list(source.fetch_area(cats, area.osm_type, area.osm_id))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Ошибка источника данных: {exc}")

    area_km2 = polygon_area_km2(area.geojson)
    summary = analytics.summarize(pois, area_km2)

    with Storage(DB_PATH) as store:
        store.save(pois)
        sid = store.record_search(
            place=req.city,
            center_lat=area.lat,
            center_lon=area.lon,
            radius_m=0.0,  # 0 = поиск по границе, а не по радиусу
            categories=[c.key for c in cats],
            result_count=len(pois),
            name=req.name,
        )

    return {
        "search_id": sid,
        "mode": "city",
        "center": [area.lat, area.lon],
        "display_name": area.display_name,
        "boundary": area.geojson,
        "area_km2": area_km2,
        "count": len(pois),
        "features": [p.as_geojson_feature() for p in pois],
        "titles": {c.key: c.title for c in categories.all_categories()},
        "summary": _summary_dict(summary),
        "rings": [],
    }


@app.post("/api/search_shape")
def api_search_shape(req: ShapeSearchRequest) -> dict:
    try:
        cats = categories.resolve(req.categories)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    geom = req.geometry or {}
    if geom.get("type") != "Polygon":
        raise HTTPException(status_code=400, detail="Ожидается GeoJSON Polygon")
    rings = geom.get("coordinates") or []
    if not rings or len(rings[0]) < 3:
        raise HTTPException(status_code=400, detail="В фигуре слишком мало точек")

    # Внешнее кольцо: GeoJSON хранит [lon, lat] → Overpass нужен (lat, lon).
    outer = rings[0]
    ring_latlon = [(pt[1], pt[0]) for pt in outer]
    cen_lat = sum(p[0] for p in ring_latlon) / len(ring_latlon)
    cen_lon = sum(p[1] for p in ring_latlon) / len(ring_latlon)

    try:
        pois = list(source.fetch_polygon(cats, ring_latlon))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Ошибка источника данных: {exc}")

    area_km2 = polygon_area_km2(geom)
    summary = analytics.summarize(pois, area_km2)

    with Storage(DB_PATH) as store:
        store.save(pois)
        sid = store.record_search(
            place=None,
            center_lat=cen_lat,
            center_lon=cen_lon,
            radius_m=0.0,
            categories=[c.key for c in cats],
            result_count=len(pois),
            name=req.name,
            geometry=json.dumps(geom),
        )

    return {
        "search_id": sid,
        "mode": "shape",
        "center": [cen_lat, cen_lon],
        "boundary": geom,
        "area_km2": area_km2,
        "count": len(pois),
        "features": [p.as_geojson_feature() for p in pois],
        "titles": {c.key: c.title for c in categories.all_categories()},
        "summary": _summary_dict(summary),
        "rings": [],
    }


EXPORT_COLUMNS = ["Название", "Категория", "Группа", "Широта", "Долгота",
                  "Адрес", "Телефон", "Часы работы", "Сайт"]


@app.post("/api/export/xlsx")
def api_export_xlsx(req: ExportRequest) -> Response:
    titles = req.titles or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты"
    ws.append(EXPORT_COLUMNS)

    for f in req.features:
        p = f.get("properties", {})
        coords = (f.get("geometry") or {}).get("coordinates") or [None, None]
        lon, lat = coords[0], coords[1]
        ws.append([
            p.get("name"),
            titles.get(p.get("category"), p.get("category")),
            p.get("group"),
            lat, lon,
            p.get("address"), p.get("phone"), p.get("opening_hours"), p.get("website"),
        ])

    # Заголовок жирным, закрепляем строку, включаем автофильтр и задаём ширины.
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{ws.max_row}"
    widths = [28, 20, 12, 11, 11, 32, 18, 22, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="infrastructure.xlsx"'},
    )


@app.get("/api/history")
def api_history() -> list[dict]:
    with Storage(DB_PATH) as store:
        return [dict(r) for r in store.list_searches()]


@app.post("/api/history/{sid}/rename")
def api_rename(sid: int, req: RenameRequest) -> dict:
    with Storage(DB_PATH) as store:
        ok = store.rename_search(sid, req.name)
    if not ok:
        raise HTTPException(status_code=404, detail="Запись истории не найдена")
    return {"ok": True}


@app.delete("/api/history/{sid}")
def api_delete(sid: int) -> dict:
    with Storage(DB_PATH) as store:
        ok = store.delete_search(sid)
    if not ok:
        raise HTTPException(status_code=404, detail="Запись истории не найдена")
    return {"ok": True}


app.mount("/static", StaticFiles(directory=str(STATIC)), name="static")


@app.get("/")
def index() -> FileResponse:
    return FileResponse(str(STATIC / "index.html"))


if __name__ == "__main__":
    import uvicorn

    # PORT задаёт большинство хостингов (Render/Railway/Fly и т.д.); локально — 8000.
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "0.0.0.0")
    print(f"Сервер запущен. Локально открой: http://127.0.0.1:{port}", file=sys.stderr)
    uvicorn.run(app, host=host, port=port, log_level="info")
